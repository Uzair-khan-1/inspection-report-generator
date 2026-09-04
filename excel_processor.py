import copy
import io
from datetime import date

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Alignment
from PIL import Image as PILImage


SHEET_NAME = "Report"

FIRST_HEADER_ROW = 8
BLOCK_SIZE = 2
EXISTING_BLOCKS = 3

OBSERVATION_COL = 1       # A
SEVERITY_COL = 2          # B
LOCATION_COL = 3          # C
FINDINGS_COL = 4          # D - Before Photo
CORRECTIVE_COL = 5        # E - After Photo
COMPLETED_COL = 6         # F
STATUS_COL = 7            # G

CENTER_WRAP = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True
)

HEADER_TEXT = {
    3: "Location / Building: ",
    4: "Findings ",
    5: "Corrective Action",
    6: "Completed On",
    7: "Status",
}


# --------------------------------------------------------------------------- #
# Geometry helpers
# --------------------------------------------------------------------------- #

def _col_width_px(width):
    """Approximate Excel column width to pixels."""
    if not width:
        width = 8.43
    return int(round(width * 7 + 5))


def _row_height_px(height):
    """Excel row height in points to pixels."""
    if not height:
        height = 15
    return int(round(height * 4 / 3))


def _stretch_to_fill(image_bytes, target_w, target_h):
    """
    Resize image exactly to target cell dimensions.
    Original aspect ratio is intentionally ignored so the
    photo completely fills the Excel cell.
    """
    target_w = max(int(target_w), 10)
    target_h = max(int(target_h), 10)

    im = PILImage.open(io.BytesIO(image_bytes))

    try:
        im = PILImage.exif_transpose(im)
    except Exception:
        pass

    if im.mode not in ("RGB", "RGBA"):
        im = im.convert("RGB")

    im = im.resize(
        (target_w, target_h),
        PILImage.LANCZOS
    )

    buf = io.BytesIO()
    im.save(buf, format="PNG")
    buf.seek(0)

    return buf


def _fill_cell_with_image(
    ws,
    image_bytes,
    col_idx,
    row_idx,
    cell_w_px,
    cell_h_px
):
    """Place photo exactly over the target Excel cell."""

    buf = _stretch_to_fill(
        image_bytes,
        cell_w_px,
        cell_h_px
    )

    marker = AnchorMarker(
        col=col_idx - 1,
        colOff=0,
        row=row_idx - 1,
        rowOff=0
    )

    size = XDRPositiveSize2D(
        pixels_to_EMU(cell_w_px),
        pixels_to_EMU(cell_h_px)
    )

    img = XLImage(buf)

    img.anchor = OneCellAnchor(
        _from=marker,
        ext=size
    )

    ws.add_image(img)


# --------------------------------------------------------------------------- #
# Row / block management
# --------------------------------------------------------------------------- #

def _copy_row_style(ws, src_row, dst_row, ncols=7):
    """Copy formatting from the template block."""

    for c in range(1, ncols + 1):
        src = ws.cell(row=src_row, column=c)
        dst = ws.cell(row=dst_row, column=c)

        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.number_format = src.number_format

    src_h = ws.row_dimensions[src_row].height

    if src_h:
        ws.row_dimensions[dst_row].height = src_h


def _add_status_conditional_formatting(ws, cell_ref):
    """Apply Open/Closed colors to Status."""

    red = PatternFill(
        start_color="FFFF0000",
        end_color="FFFF0000",
        fill_type="solid"
    )

    green = PatternFill(
        start_color="FF00B050",
        end_color="FF00B050",
        fill_type="solid"
    )

    ws.conditional_formatting.add(
        cell_ref,
        FormulaRule(
            formula=[
                f'NOT(ISERROR(SEARCH("OPEN",{cell_ref})))'
            ],
            fill=red
        )
    )

    ws.conditional_formatting.add(
        cell_ref,
        FormulaRule(
            formula=[
                f'NOT(ISERROR(SEARCH("CLOSED",{cell_ref})))'
            ],
            fill=green
        )
    )


def _set_severity_cell(ws, cell_ref, severity):
    """
    Put severity directly into the correct B-cell and apply
    its color. This guarantees that severity is visible in Excel.
    """

    severity = (severity or "Low").strip().title()

    cell = ws[cell_ref]
    cell.value = severity
    cell.alignment = CENTER_WRAP

    # Direct cell colors
    if severity == "High":
        cell.fill = PatternFill(
            start_color="FFFF0000",
            end_color="FFFF0000",
            fill_type="solid"
        )

    elif severity == "Medium":
        cell.fill = PatternFill(
            start_color="FFFFC000",
            end_color="FFFFC000",
            fill_type="solid"
        )

    else:
        cell.fill = PatternFill(
            start_color="FF00B050",
            end_color="FF00B050",
            fill_type="solid"
        )


def _ensure_block(ws, index):
    """
    Return header row and data row.

    Observation 1:
        Header = 8
        Data   = 9

    Observation 2:
        Header = 10
        Data   = 11

    Observation 3:
        Header = 12
        Data   = 13
    """

    header_row = FIRST_HEADER_ROW + index * BLOCK_SIZE
    data_row = header_row + 1

    if index < EXISTING_BLOCKS:
        return header_row, data_row

    # Insert new 2-row block after existing blocks
    ws.insert_rows(header_row, amount=BLOCK_SIZE)

    _copy_row_style(
        ws,
        FIRST_HEADER_ROW,
        header_row
    )

    _copy_row_style(
        ws,
        FIRST_HEADER_ROW + 1,
        data_row
    )

    for col, text in HEADER_TEXT.items():
        ws.cell(
            row=header_row,
            column=col,
            value=text
        )

    _add_status_conditional_formatting(
        ws,
        f"G{data_row}"
    )

    return header_row, data_row


def _ensure_equal_photo_columns(ws):
    """Make columns D and E equal width."""

    d_dim = ws.column_dimensions[
        get_column_letter(FINDINGS_COL)
    ]

    e_dim = ws.column_dimensions[
        get_column_letter(CORRECTIVE_COL)
    ]

    if d_dim.width:
        e_dim.width = d_dim.width


# --------------------------------------------------------------------------- #
# Main report generator
# --------------------------------------------------------------------------- #

def generate_report(items, template_path, output_path):

    wb = load_workbook(template_path)
    ws = wb[SHEET_NAME]

    # --------------------------------------------------------- #
    # Report heading
    # --------------------------------------------------------- #

    ws["A7"] = "FADHILI INSPECTION REPORT"

    # Keep heading formatting from existing template
    ws["A7"].alignment = copy.copy(
        ws["A7"].alignment
    )

    # --------------------------------------------------------- #
    # Make photo columns equal
    # --------------------------------------------------------- #

    _ensure_equal_photo_columns(ws)

    row_h = _row_height_px(
        ws.row_dimensions[
            FIRST_HEADER_ROW + 1
        ].height
    )

    d_w = _col_width_px(
        ws.column_dimensions[
            get_column_letter(FINDINGS_COL)
        ].width
    )

    e_w = _col_width_px(
        ws.column_dimensions[
            get_column_letter(CORRECTIVE_COL)
        ].width
    )

    # --------------------------------------------------------- #
    # Add observations
    # --------------------------------------------------------- #

    for i, item in enumerate(items):

        header_row, data_row = _ensure_block(
            ws,
            i
        )

        # ----------------------------------------------------- #
        # A - Observation Number
        # ----------------------------------------------------- #

        obs_cell = ws.cell(
            row=data_row,
            column=OBSERVATION_COL
        )

        obs_cell.value = i + 1
        obs_cell.alignment = CENTER_WRAP

        # ----------------------------------------------------- #
        # B - Category of Severity
        # ----------------------------------------------------- #

        severity = item.get("severity") or "Low"

        _set_severity_cell(
            ws,
            f"B{data_row}",
            severity
        )

        # ----------------------------------------------------- #
        # C - Location / Building + Finding
        # ----------------------------------------------------- #

        location = item.get("location", "")
        finding = item.get("finding", "")

        location_finding = (
            f"{location}\n\n{finding}"
        ).strip()

        location_cell = ws.cell(
            row=data_row,
            column=LOCATION_COL
        )

        location_cell.value = location_finding

        location_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        # ----------------------------------------------------- #
        # F - Completed On
        # ----------------------------------------------------- #

        completed_cell = ws.cell(
            row=data_row,
            column=COMPLETED_COL
        )

        completed_cell.value = (
            item.get("completed_on")
            or date.today()
        )

        completed_cell.alignment = CENTER_WRAP

        # ----------------------------------------------------- #
        # G - Status
        # ----------------------------------------------------- #

        status_cell = ws.cell(
            row=data_row,
            column=STATUS_COL
        )

        status_cell.value = (
            item.get("status")
            or "Open"
        )

        status_cell.alignment = CENTER_WRAP

        # ----------------------------------------------------- #
        # Before Photo -> D
        # ----------------------------------------------------- #

        if item.get("before_bytes"):

            _fill_cell_with_image(
                ws,
                item["before_bytes"],
                col_idx=FINDINGS_COL,
                row_idx=data_row,
                cell_w_px=d_w,
                cell_h_px=row_h
            )

        # ----------------------------------------------------- #
        # After Photo -> E
        # ----------------------------------------------------- #

        if item.get("after_bytes"):

            _fill_cell_with_image(
                ws,
                item["after_bytes"],
                col_idx=CORRECTIVE_COL,
                row_idx=data_row,
                cell_w_px=e_w,
                cell_h_px=row_h
            )

    # --------------------------------------------------------- #
    # Save report
    # --------------------------------------------------------- #

    wb.save(output_path)

    return output_path
